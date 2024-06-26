action=input("Seleccione una opción (inserte solo el número): \n1. Cargar un pedido de CAF para generar Excel legible. \n2. Comparar dos versiones de un pedido de CAF \n3. Cargar un CBC para generar un excel temporal con la propuesta de autorrellenado. \n4. Cargar un excel temporal para autorellenar un CbC \n")
'''
action=input("Seleccione una opción (inserte solo el número): \n1. Crear una Base de Datos (SOLO ADMIN). \n2. Cargar un CBC a la Base de Datos (SOLO ADMIN). \n3. Cargar un fichero temporal a la base de datos (SOLO ADMIN). \n4. Cargar un CBC para generar un excel temporal con la propuesta de autorrellenado. \n5. Cargar un excel temporal para autorellenar un CbC \n")
if action=="1":
    pwd=getpass.getpass("INTRODUZCA LA CONTRASEÑA: ")
    if pwd=="MCG@Trenes":
        import createDatabase
    else:
        messagebox.showinfo("Acceso Denegado","La contraseña introducida es incorrecta.")
'''

import pandas as pd








######## FUNCIÓN PARA CARGAR UN CBC Y AUTORELLENAR LAS CLÁUSULAS EN BASE A LO QUE HAYA EN BASE DE DATOS, GUARDANDO LOS RESULTADOS EN UN EXCEL TEMPORAL


'''
#Ejemplo: python LoadDatabase.py 'C:/Users/mcuberos/Desktop/AppGestorRequisitos_old/Python/D0000016800_EEFAE SALOON HVAC P2_Ed-HC_230517.xlsx' '7' 'G'''

import os
import re
import pandas as pd
from tkinter import messagebox
import sys
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfilename 
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import pyodbc 
import json



dfhead = pd.DataFrame({'OrderNumber': [],
                   'Version': [],
                   'OrderDate': [],
                   'Reference':[]})

dfdetail = pd.DataFrame({'ItemNumber': [],
                    'MaterialCode': [],
                    'Description': [],
                    'MaterialGroup':[],
                    'Contract':[],
                    'Deadline':[],
                    'Quantity':[],
                    'UnitPrice':[],
                    'Discount':[],
                    'TotalValue':[],
                    'Quantity_2':[],
                    'Deadline_2':[],
                    'QuantityDelivered':[],
                    'QuantityReceived':[],
                    'QuantityPending':[],
                    'DeliveryAddress':[]})

'''
dftemp = pd.DataFrame({'Id_Req': [],
                   'Desc_Req': [],
                   'ID_Req_BBDD': [],
                   'Descr_Req_BBDD':[],
                   '%Coincidencia':[],
                   'Resp_BBDD':[],
                   'Coment_BBDD':[],
                   'Nueva_Respuesta':[],
                   'Nuevos_Comentarios':[],
                   'Proyecto_Origen':[],
                   'Fichero_Origen':[],
                   'Tipo_Vehiculo':[],
                   'Tipo_Entregable':[]})
'''

def leer_configuracion(fichero_config):
    with open(fichero_config, 'r') as archivo:
        configuracion = json.load(archivo)
    return configuracion

if getattr(sys,'frozen',False):
    PathName=os.path.dirname(sys.executable)
else:
    PathName=os.path.dirname(os.path.abspath(__file__))

#PathName=os.path.dirname(__file__)
fileConfig = PathName + '/config.json'
#fileConfig="C:/AppGestorRequisitos/config.json"

configuracion=leer_configuracion(fileConfig)

# Acceder a los datos
print("Creando conexión con la base de datos...")
direccion_bd = configuracion['database']['host']
servidor=direccion_bd.replace("/","\\")
database=configuracion['database']['database_name']
username=configuracion['database']['username']
tablename=configuracion['database']['table_name']
password='Prueb@sManuel23'
# Cadena de conexión
connection_string = f'DRIVER={{SQL Server}};SERVER={servidor};DATABASE={database};UID={username};PWD={password}'
# Establecer la conexión
connection = pyodbc.connect(connection_string)
cursor=connection.cursor()
cursor.execute("SELECT * FROM T_REQUISITOS")
listadoRequisitos=cursor.fetchall()
cursor.close()
connection.close()

def AddClauseToTemp(clausula):
    """Recorre la bbdd buscando la cláusula para informar el excel original"""
    global dftemp
    global connection_string
    global listadoRequisitos
    '''
    # Establecer la conexión
    connection = pyodbc.connect(connection_string)
    cursor=connection.cursor()
    cursor.execute("SELECT * FROM REQUISITOS")
    listadoRequisitos=cursor.fetchall()
    '''

    ExisteClausula=FALSE
    best_accuracy=0
    for requisito in listadoRequisitos:
        if clausula[1]==requisito[1] or clausula[2]==requisito[2]:
            #EXISTE LA CLAUSULA TAL CUAL EN BASE DE DATOS. SE DEBE TOMAR ESE VALOR
            ExisteClausula=TRUE
            if requisito[4]!="":
                nuevo_comentario="V0 " + requisito[10] + ":\n" + requisito[4]
            else:
                nuevo_comentario=requisito[4]
            nueva_fila = pd.Series([clausula[1], clausula[2], requisito[1],requisito[2],100,requisito[3],requisito[4],requisito[3],nuevo_comentario,requisito[10],requisito[11],requisito[5],requisito[12]], index=dftemp.columns)
            dftemp = dftemp._append(nueva_fila, ignore_index=True)
        else:
            accuracy=CheckClause(clausula[2],requisito[2])
            if accuracy>best_accuracy:
                if requisito[4]!="":
                    nuevo_comentario="V0 " + requisito[10] + ":\n" + requisito[4]
                else:
                    nuevo_comentario=requisito[4]
                nueva_fila = pd.Series([clausula[1], clausula[2], requisito[1],requisito[2],accuracy,requisito[3],requisito[4],requisito[3],nuevo_comentario,requisito[10],requisito[11],requisito[5],requisito[12]], index=dftemp.columns)
                best_accuracy=accuracy
                if accuracy>80: #o meto este IF o meto la linea 59, para guardar solo la mejor coincidencia
                    dftemp = dftemp._append(nueva_fila, ignore_index=True) #SI METO ESTA LÍNEA, VOY A AÑADIR EN EL EXCEL TEMPORAL TODAS LAS COINCIDENCIAS QUE ENCUENTRE PARA LA CLAUSULA CON ACCURACY >80
        
    if ExisteClausula==FALSE:
        #SI NO SE HA ENCONTRADO EXACTAMENTE LA CLAUSULA, AÑADO NUEVA_FILA, QUE ES EL REQUISITO QUE HA ENCONTRADO CON MAYOR ACCURACY.
        if best_accuracy==0:
            nueva_fila = pd.Series([clausula[1], clausula[2], "NO HAY COINCIDENCIAS","NO HAY COINCIDENCIAS",accuracy,"NO HAY COINCIDENCIAS","NO HAY COINCIDENCIAS","","","","","",""], index=dftemp.columns)
        if best_accuracy<80:
            dftemp = dftemp._append(nueva_fila, ignore_index=True)
    
    return dftemp


def CheckClause(newClause,requirement):
    """ESTA FUNCIÓN COMPRUEBA SI UN REQUISITO NUEVO ES IGUAL A OTRO GUARDADO EN BBDD.
        DEVUELVE UN % DE COINDICENCIA ENTRE LOS DOS REQUISITOS"""
    if newClause==requirement:
        accuracy=100
    else:
        accuracy=0
        longParcial=30
        numTramos=int(len(newClause)/longParcial)+1
        #longParcial=int(len(newClause)/15)
        for aux in range(numTramos):
            cadena=newClause[aux*longParcial:longParcial*(aux+1)]
            if re.search(re.escape(cadena),requirement):
                accuracy=accuracy+99/numTramos

    return accuracy
print("SELECCIONE EL FICHERO DE CBC A INFORMAR ")
fileName=askopenfilename()
filaHeader=input("INDIQUE LA FILA DONDE SE ENCUENTRA LA CABECERA DEL CBC ")
colIdReq=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LOS IDs DEL REQUISITO (A,B,C,D,...) ")
colClause=input("INDIQUE LA COLUMNA DONDE SE ENCUENTRAN LAS DESCRIPCIONES DE LOS REQUISITOS (A,B,C,D,...) ") #para convertir la columna (letra) a número: ord(colClause.lower())-96
nombre_hoja=input("INDIQUE EL NOMBRE DE LA HOJA DONDE SE ENCUENTRAN LOS REQUISITOS (POR DEFECTO, Requirements)")
tipo_vehiculo=input("INDIQUE EL TIPO DE VEHÍCULO DEL CBC (TRANVÍA, METRO, ...) ")
entregable_cbc=input("INDIQUE EL TIPO DE EQUIPO AL QUE HACE REFERENCIA EL CBC (SALA, CABINA, GENERAL): ")
#proy_origen=input("INDIQUE EL NOMBRE DEL PROYECTO: ")
#fichero_origen=os.path.split(fileName)[1]


#GENERO EL DATAFRAME OMITIENDO LAS FILAS QUE SE CORRESPONDEN A TÍTULOS QUE TIENEN EL FONDO GRIS 
# PRIMERO, GUARDO EN UNA TUPLA TODAS LAS FILAS DE TÍTULOS  
print("Generando dataframe con todas las cláusulas del CbC...")
book=load_workbook(fileName)
if nombre_hoja=="":
    nombre_hoja="Requirements"

hoja_excel=book[nombre_hoja]
fila=int(filaHeader)+1
celda=hoja_excel[colIdReq+str(fila)]
filas_titulos=[]
#TENGO QUE MODIFICAR ALGO EN EL WHILE, PORQUE CUANDO ENCUENTRA UNA LÍNEA EN BLANCO DEJA DE RECORRERLO, Y PUEDE HABERLA. tengo que jugar con el tamaño del df global, contando titulos
df=pd.read_excel(fileName, sheet_name=nombre_hoja,header=int(filaHeader)-1,keep_default_na=FALSE)

print("Descartando filas que hacen referencia a títulos de apartados en el CbC......")
for aux in range(len(df)):
    if (celda.fill.fgColor.rgb!="00000000" and celda.fill.fgColor.type!="theme") or celda.value==None: #busco las celdas que tengan relleno distinto de vacío y de blanco
        filas_titulos.append(fila-1) #guardo fila-1 porque el dataframe trabaja con la fila 0
    fila+=1
    celda=hoja_excel["A"+str(fila)]

df=pd.read_excel(fileName, sheet_name=nombre_hoja,header=int(filaHeader)-1,skiprows=filas_titulos,keep_default_na=FALSE)

#RECORRO EL DATAFRAME EN LA COLUMNA colClause 
print("Buscando respuestas en cláusulas similares de la base de datos...")
for kk in range(len(df)):
    if kk%10==0:
        print(str(kk) + "/" + str(len(df)))
    if len(df.iloc[kk][(ord(colClause.lower())-97)])>5:
        #defino la variable clausula como una tupla que contiene: id_requisito, descripción de la clausula, la respuesta, comentarios
        clausula=(kk,df.iloc[kk][(ord(colIdReq.lower())-97)],df.iloc[kk][(ord(colClause.lower())-97)])
        AddClauseToTemp(clausula)


if len(dftemp)>0:
    print("Creando excel CbC Temporal con las respuestas encontradas...")
    Ruta=os.path.dirname(fileName) #busco la ruta del CBC para crear ahí el cbctemp
    excelTemp=(Ruta + "/CBCTemporal.xlsx")
    writer = pd.ExcelWriter(excelTemp,engine='xlsxwriter')
    dftemp.to_excel(writer, 'CBC Temporal', index=False)
    writer.close()
    #DOY FORMATO A COLUMNAS 
    libro=load_workbook(excelTemp)
    hoja=libro.active
    columnas_a_formatear = ["H","I"]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in columnas_a_formatear:
        for celda in hoja[col]:
            celda.fill=fill


    filas_borrar=[]

    for i in range(len(hoja["A"])):
        if hoja["A"+str(i+1)].value==hoja["A"+str(i+2)].value:
            k=2
            while hoja["A"+str(i+k)].value==hoja["A"+str(i+1)].value:
                if hoja["L"+str(i+1)].value==tipo_vehiculo and hoja["M"+str(i+1)].value==entregable_cbc and (k+i) not in filas_borrar: #si en la primera fila coincide tipo veh y entregable, borro la otra fila
                    filas_borrar.append(i+k)
                elif hoja["L"+str(i+k)].value==tipo_vehiculo and hoja["M"+str(i+k)].value==entregable_cbc and (i+1) not in filas_borrar: #y viceversa
                    filas_borrar.append(i+1)
                else:
                    hoja["A"+str(i+1)].fill=fill
                    hoja["A"+str(i+k)].fill=fill
                k=k+1
            
    for f in filas_borrar[::-1]:
        print(f)
        hoja.delete_rows(f)


    fillHeader = PatternFill(start_color="EF9191", end_color="EF9191", fill_type="solid")
    font = Font(bold=True)
    encabezado=hoja[1]
    for celda in encabezado:
        celda.fill=fillHeader
        celda.font=font

    libro.save(excelTemp)

    messagebox.showinfo("EXCEL TEMPORAL CREADO","SE HA CREADO UN EXCEL TEMPORAL CON LAS COINCIDENCIAS DE LAS CLÁUSULAS ENCONTRADAS EN LA RUTA " + Ruta)   


messagebox.showinfo("FIN PROCESO","EJECUCIÓN FINALIZADA CON ÉXITO")